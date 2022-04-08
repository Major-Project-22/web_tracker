# -*- coding: utf-8 -*-
"""
Created on Thu Mar 18 12:36:09 2021

@author: nobin
"""
import pymongo
import openpyxl
from flask import Flask,request,jsonify
from flask_cors import CORS
client=pymongo.MongoClient('mongodb+srv://appDB:Banglore1@cluster0.opser.mongodb.net/myFirstDatabase?retryWrites=true&w=majority')
Doctor=client.Doctor_Database
Nurse=client.Nurse_Database
Patient=client.Patient_Database
Log=client.Log_Database
Record=client.Record_Database
Stock=client.Stock_Database
Stock_table=Stock.Table
Record_table=Record.Table
Log_table=Log.Table
Doctor_table=Doctor.Table
Nurse_table=Nurse.Table
Patient_table=Patient.Table
page=Flask(__name__)
CORS(page)
Ward=""
nurseWard=""
@page.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin','*')
    response.headers.add('Access-Control-Allow-Headers','Content-Type')
    return response
@page.route('/Aquantity',methods=['GET'])
def Aquantity():
    d=Stock_table.find_one({'Medicine_name':'Medicine 1'})
    return jsonify({'quantity':d['Quantity']})
@page.route('/Bquantity',methods=['GET'])
def Bquantity():
    d=Stock_table.find_one({'Medicine_name':'Medicine 1'})
    return jsonify({'quantity':d['Quantity']})
@page.route('/Cquantity',methods=['GET'])
def Cquantity():
    d=Stock_table.find_one({'Medicine_name':'Medicine 3'})
    return jsonify({'quantity':d['Quantity']})
@page.route('/Dquantity',methods=['GET'])
def Dquantity():
    d=Stock_table.find_one({'Medicine_name':'Medicine 4'})
    return jsonify({'quantity':d['Quantity']})
@page.route('/fetchA',methods=['GET'])
def fetchA():
   if Log_table.count_documents({'Medication':'Medicine 1'})==0:
        return jsonify({})
   else:
        d=[]
        for documents in Log_table.find({'Medication':'Medicine 1'}):
            d.append({'Medication':documents['Medication'],'Patient':documents['Patient'],'Time':documents['Time']})
        return jsonify(d)
@page.route('/fetchB',methods=['GET'])
def fetchB():
   if Log_table.count_documents({'Medication':'Medicine 2'})==0:
        return jsonify({})
   else:
        d=[]
        for documents in Log_table.find({'Medication':'Medicine 2'}):
            d.append({'Medication':documents['Medication'],'Patient':documents['Patient'],'Time':documents['Time']})
        return jsonify(d)
@page.route('/fetchC',methods=['GET'])
def fetchC():
   if Log_table.count_documents({'Medication':'Medicine 3'})==0:
        return jsonify({})
   else:
        d=[]
        for documents in Log_table.find({'Medication':'Medicine 3'}):
            d.append({'Medication':documents['Medication'],'Patient':documents['Patient'],'Time':documents['Time']})
        return jsonify(d)
@page.route('/fetchD',methods=['GET'])
def fetchD():
   if Log_table.count_documents({'Medication':'Medicine 4'})==0:
        return jsonify({})
   else:
        d=[]
        for documents in Log_table.find({'Medication':'Medicine 4'}):
            d.append({'Medication':documents['Medication'],'Patient':documents['Patient'],'Time':documents['Time']})
        return jsonify(d)
@page.route('/checkStock',methods=['GET'])
def checkStock():
    if Log_table.count_documents({})==0:
        return jsonify({})
    else:
        d=[]
        for documents in Log_table.find({}):
            d.append({'Medication':documents['Medication'],'Patient':documents['Patient'],'Time':documents['Time']})
        return jsonify(d)
@page.route('/restock',methods=['GET','POST'])
def restock():
    if request.method=="POST":
        Stock_table.delete_one({'Medicine_name':request.json['Medicine_name']})
        Stock_table.insert_one({'Medicine_name':request.json['Medicine_name'],'Quantity':10})
        return jsonify({})
@page.route('/nurselogin',methods=['POST','GET'])
def nurselogin():
    global nurseWard
    if request.method=="POST":
        d=Nurse_table.find_one({'Nurse_id':request.json['ID'],'Name':request.json['Name']})
        if d==None:
            return jsonify({})
        else:
            nurseWard=d['Ward']
            return jsonify({'Correct':True})
@page.route('/nursepatients',methods=['GET','POST'])
def nursepatients():
    global nurseWard
    d=[]
    for documents in Patient_table.find({'Ward':nurseWard}):
        d.append({'Name':documents["Name"],'Patient_ID':documents['Patient_id'],'Medicine_1':documents['Medicine_1'],'Medicine_2':documents['Medicine_2'],'Medicine_3':documents['Medicine_3'],'Medicine_4':documents['Medicine_4']})
    dd=sorted(d,key=lambda x:int(x['Patient_ID'][2:]))
    return jsonify(dd)
@page.route('/doctorverify',methods=["POST","GET"])
def doctorverify():
    global Ward
    if request.method=="POST":
        d=Doctor_table.find_one({'Doctor_id':request.json['ID'],'Name':request.json['Name']})
        if d == None:
            print("hit")
            return jsonify({})
        else:
            Ward=d['Ward']
            return jsonify({'Name':d['Name'],'ID':d['Doctor_id'],'ward':d['Ward']})
@page.route('/doctorpatients',methods=["GET"])
def doctorpatients():
    global Ward
    d=[]
    for documents in Patient_table.find({'Ward':Ward}):
        d.append({'Name':documents["Name"],'Patient_ID':documents['Patient_id'],'Medicine_1':documents['Medicine_1'],'Medicine_2':documents['Medicine_2'],'Medicine_3':documents['Medicine_3'],'Medicine_4':documents['Medicine_4']})
    dd=sorted(d,key=lambda x:int(x['Patient_ID'][2:]))
    return jsonify(dd)
@page.route('/fetchpatient',methods=['GET','POST'])
def fetchpatient():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    return jsonify({'Name':d['Name'],'ID':d['Patient_id'],'ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':d['Medicine_2'],'Medicine_3':d['Medicine_3'],'Medicine_4':d['Medicine_4']})
@page.route('/fetchlog',methods=['GET','POST'])
def fetchlog():
    d=[]
    print(request)
    if Record_table.count_documents({'Patient_id':request.json['ID']})==0:
        return jsonify({'records_exist':False})
    else:
        for document in Record_table.find({'Patient_id':request.json['ID']}):
            d.append({'Doctor':Doctor_table.find_one({'Doctor_id':document['Doctor_id']})['Name'],'Record':document['Record'],'Datetime':document['Datetime']})
        return jsonify(d)
@page.route('/submitrecord',methods=['GET','POST'])
def submitrecord():
    Record_table.insert_one(request.json)
    print("record Added")
    return jsonify({})
@page.route('/removemedicine1',methods=['GET','POST'])
def removemedicine1():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':False,'Medicine_2':d['Medicine_2'],'Medicine_3':d['Medicine_3'],'Medicine_4':d['Medicine_4']})
    return jsonify({})
@page.route('/removemedicine2',methods=['GET','POST'])
def removemedicine2():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':False,'Medicine_3':d['Medicine_3'],'Medicine_4':d['Medicine_4']})
    return jsonify({})
@page.route('/removemedicine3',methods=['GET','POST'])
def removemedicine3():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':d['Medicine_2'],'Medicine_3':False,'Medicine_4':d['Medicine_4']})
    return jsonify({})
@page.route('/removemedicine4',methods=['GET','POST'])
def removemedicine4():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':d['Medicine_2'],'Medicine_3':d['Medicine_3'],'Medicine_4':False})
    return jsonify({})
@page.route('/addmedicine1',methods=['GET','POST'])
def addmedicine1():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':True,'Medicine_2':d['Medicine_2'],'Medicine_3':d['Medicine_3'],'Medicine_4':d['Medicine_4']})
    return jsonify({})
@page.route('/addmedicine2',methods=['GET','POST'])
def addmedicine2():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':True,'Medicine_3':d['Medicine_3'],'Medicine_4':d['Medicine_4']})
    return jsonify({})
@page.route('/addmedicine3',methods=['GET','POST'])
def addmedicine3():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':d['Medicine_2'],'Medicine_3':True,'Medicine_4':d['Medicine_4']})
    return jsonify({})
@page.route('/addmedicine4',methods=['GET','POST'])
def addmedicine4():
    d=Patient_table.find_one({'Patient_id':request.json['ID']})
    Patient_table.delete_one({'Patient_id':request.json['ID']})
    Patient_table.insert_one({'Patient_id':request.json['ID'],'Name':d['Name'],'Ward':d['Ward'],'Medicine_1':d['Medicine_1'],'Medicine_2':d['Medicine_2'],'Medicine_3':d['Medicine_3'],'Medicine_4':True})
    return jsonify({})


# class Dbinit():
#     def __init__(self):
#         dbDocument=openpyxl.load_workbook("Medicines.xlsx")
#         Medicine_price=dbDocument['Sheet1']
#         Doctor_list=dbDocument['Sheet2']
#         Patient_list=dbDocument['Sheet3']
#         Nurse_list=dbDocument['Sheet4']
#         Stock_list=dbDocument['Sheet5']
        # medicine_price_table.drop()
        # Doctor_table.drop()
        # Nurse_table.drop()
        # Patient_table.drop()
        # Stock_table.drop()
      
        # for rows in range(2,Doctor_list.max_row+1):
        #     table_dict={'Doctor_id':'Dr'+str(Patient_list.cell(rows,1).value),'Name':Doctor_list.cell(rows,2).value,'Ward':Doctor_list.cell(rows,3).value}
        #     Doctor_table.insert_one(table_dict)
        # for rows in range(2,Nurse_list.max_row+1):
        #     table_dict={'Nurse_id':'Nr'+str(Nurse_list.cell(rows,1).value),'Name':Nurse_list.cell(rows,2).value,'Ward':Nurse_list.cell(rows,3).value}
        #     Nurse_table.insert_one(table_dict)
#         for rows in range(2,Patient_list.max_row+1):
#             table_dict={'Patient_id':'Pt'+str(Patient_list.cell(rows,1).value),'Name':Patient_list.cell(rows,2).value,'Ward':Patient_list.cell(rows,3).value,'Medicine_1':False,'Medicine_2':False,'Medicine_3':False,'Medicine_4':False}
#             Patient_table.insert_one(table_dict)
# Dbinit()
if __name__ == '__main__':
    page.run(debug=True,port=5000)